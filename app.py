import chainlit as cl
from PyPDF2 import PdfReader
from langchain.text_splitter import RecursiveCharacterTextSplitter,CharacterTextSplitter
from langchain.vectorstores import FAISS
from langchain.embeddings import OpenAIEmbeddings,HuggingFaceInstructEmbeddings
from langchain.chains import ConversationalRetrievalChain
from langchain.chat_models import ChatOpenAI
from langchain.memory import ConversationBufferMemory
from langchain.document_loaders import WebBaseLoader,UnstructuredURLLoader, PyPDFLoader, CSVLoader,JSONLoader,UnstructuredMarkdownLoader
import os,pickle
from langchain import OpenAI
from bs4 import BeautifulSoup
import json,markdown,csv,io,textract
from pathlib import Path
from openpyxl import load_workbook

welcome_message = """### Welcome to Chat with 👋
You can upload a file or paste a url to chat with.
I support many types of files and urls.
"""


supported_file_types = ["pdf","html","plain","json","markdown","python-script"]

embeddings = OpenAIEmbeddings()
memory = ConversationBufferMemory(
    memory_key="chat_history", return_messages=True
)
llm = ChatOpenAI(temperature=0, model="gpt-3.5-turbo")
condense_question_llm = ChatOpenAI(temperature=0, model='gpt-3.5-turbo')

text_splitter = RecursiveCharacterTextSplitter(
    chunk_size=2500, chunk_overlap=200)

def get_files_text(file_response):
    if file_response is None:
        return ""

    text = ""

    # Access the content and name properties directly
    file_content = file_response.content
    file_name = file_response.name

    if file_content is None:
        return ""

    if file_name.endswith('.pdf'):
        pdf_reader = PdfReader(io.BytesIO(file_content))
        for page in pdf_reader.pages:
            text += page.extract_text()
    elif file_name.endswith('.txt'):
        text += file_content.decode('utf-8')
    elif file_name.endswith('.md'):
        md_content = file_content.decode('utf-8')
        html_content = markdown.markdown(md_content)
        soup = BeautifulSoup(html_content, 'html.parser')
        text += soup.get_text()
    elif file_name.endswith('.json'):
        json_content = json.loads(file_content)
        if isinstance(json_content, dict):
            text += ' '.join(str(value) for value in json_content.values())
        elif isinstance(json_content, list):
            text += ' '.join(str(item) for item in json_content)
    elif file_name.endswith('.csv'):
        csv_content = file_content.decode('utf-8')
        csv_file = io.StringIO(csv_content)
        csv_reader = csv.reader(csv_file)
        for row in csv_reader:
            text += ' '.join(row) + ' '
    elif file_name.endswith('.xlsx'):
        workbook = load_workbook(io.BytesIO(file_content), read_only=True)
        for sheet in workbook:
            for row in sheet.iter_rows():
                for cell in row:
                    text += str(cell.value) + ' '
    elif file_name.endswith('.py'):
        text += file_content.decode('utf-8')

    return text

def get_text_chunks(text):
    text_splitter = CharacterTextSplitter(
        separator="\n",
        chunk_size=2000,
        chunk_overlap=200,
        length_function=len
    )
    chunks = text_splitter.split_text(text)
    return chunks
def get_text_chunks1(text):
    text_splitter = CharacterTextSplitter(
        separator="\n",
        chunk_size=500,
        chunk_overlap=200,
    )
    chunks = text_splitter.split_documents(text)
    return chunks
    

def get_vectorstore(text_chunks):
    embeddings = OpenAIEmbeddings()
    # embeddings = HuggingFaceInstructEmbeddings(model_name="hkunlp/instructor-xl")
    vectorstore = FAISS.from_texts(texts=text_chunks, embedding=embeddings)
    return vectorstore

def get_vectorstore1(text_chunks):
     embeddings = OpenAIEmbeddings()
     #embeddings = HuggingFaceInstructEmbeddings(model_name="hkunlp/instructor-xl")
     vectorstore = FAISS.from_documents(documents=text_chunks, embedding=embeddings)
     return vectorstore

def get_conversation_chain(vectorstore):
    llm = ChatOpenAI()
    #llm = HuggingFaceHub(repo_id="google/flan-t5-xxl", model_kwargs={"temperature":0.5, "max_length":512})

    memory = ConversationBufferMemory(
        memory_key='chat_history', return_messages=True)
    conversation_chain = ConversationalRetrievalChain.from_llm(
        llm=llm,
        retriever=vectorstore.as_retriever(),
        memory=memory
    )
    return conversation_chain


@cl.langchain_factory
def main():
    cl.Message(content=welcome_message).send()

    cleaned_resp = None
    valid_responses = ["file", "url"]
    
    while cleaned_resp not in valid_responses:
        resp = cl.AskUserMessage(
            "What would you like to chat with, a File or a URL?"
        ).send()
        if 'content' in resp:
            cleaned_resp = resp['content'].strip().lower()

    if cleaned_resp == "file":
        file = None
        while file is None:
            file = cl.AskFileMessage(
                content='Upload your file:', accept=supported_file_types, timeout=180, max_size_mb=500,
            ).send()

        raw_text = get_files_text(file)

        # get the text chunks
        text_chunks = get_text_chunks(raw_text)

        # create vector store
        vectorstore = get_vectorstore(text_chunks)

        chain = get_conversation_chain(vectorstore)

        cl.Message(
            content=f"`{file.name}` uploaded! How can I help you?"
        ).send()

        return chain

    if cleaned_resp == "url":
        url = None
        while url is None:
            url = cl.AskUserMessage(
                content="Paste Your URL:"
            ).send()
            url = url['content']
        
        print(type(url))
        
        loaders=WebBaseLoader(url)

        data = loaders.load()
        print(data)
        docs = get_text_chunks1(data)
        vectorstore = get_vectorstore1(docs)
        chain = get_conversation_chain(vectorstore)

        cl.Message(
            content=f"URL uploaded! How can I help you?"
        ).send()

        return chain
